import io
import logging
import secrets
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, Request, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, EmailStr
from sqlalchemy import func
from sqlalchemy.orm import Session
from app.api.deps import get_current_admin
from app.core.database import get_db
from app.core.security import hash_password, verify_password, create_access_token
from app.models.address import Address
from app.models.order import Order
from app.models.user import User
from app.schemas.auth import RegisterRequest, LoginRequest, TokenResponse
from app.services.email import send_verification_email

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/auth", tags=["auth"])


def _new_verification_token() -> str:
    return secrets.token_urlsafe(32)


@router.post("/register", response_model=TokenResponse, status_code=status.HTTP_201_CREATED)
def register(request: Request, body: RegisterRequest, db: Session = Depends(get_db)):
    if db.query(User).filter(User.email == body.email).first():
        raise HTTPException(status_code=400, detail="Email already registered")
    if body.phone and db.query(User).filter(User.phone == body.phone).first():
        raise HTTPException(status_code=400, detail="Phone number already registered")

    token = _new_verification_token()
    user = User(
        name=body.name,
        email=body.email,
        phone=body.phone,
        password_hash=hash_password(body.password),
        email_verified=False,
        verification_token=token,
    )
    db.add(user)
    db.commit()
    db.refresh(user)

    try:
        send_verification_email(user, token)
    except Exception as e:  # never fail registration because email failed
        logger.error("Verification email failed for %s: %s", user.email, e)

    access = create_access_token({"sub": str(user.id)})
    return TokenResponse(access_token=access, user_id=user.id, name=user.name, is_admin=user.is_admin)


@router.post("/login", response_model=TokenResponse)
def login(request: Request, body: LoginRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.email == body.email).first()
    if not user or not verify_password(body.password, user.password_hash):
        logger.warning("Failed login attempt for %s from %s", body.email, request.client.host)
        raise HTTPException(status_code=401, detail="Invalid email or password")

    logger.info("User %s logged in from %s", user.id, request.client.host)
    token = create_access_token({"sub": str(user.id)})
    return TokenResponse(access_token=token, user_id=user.id, name=user.name, is_admin=user.is_admin)


class VerifyResponse(BaseModel):
    verified: bool
    email: str | None = None


@router.get("/verify-email", response_model=VerifyResponse)
def verify_email(token: str, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.verification_token == token).first()
    if not user:
        raise HTTPException(status_code=400, detail="Invalid or expired verification link")
    user.email_verified = True
    user.verification_token = None
    db.commit()
    logger.info("Email verified for %s", user.email)
    return VerifyResponse(verified=True, email=user.email)


class ResendRequest(BaseModel):
    email: EmailStr


@router.post("/resend-verification")
def resend_verification(body: ResendRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.email == body.email).first()
    # Do not reveal whether the account exists
    if user and not user.email_verified:
        token = _new_verification_token()
        user.verification_token = token
        db.commit()
        send_verification_email(user, token)
    return {"ok": True}


def _user_rows(db: Session):
    """Aggregate every user with their addresses, order stats, and latest order."""
    order_stats = (
        db.query(
            Order.user_id,
            func.count(Order.id).label("order_count"),
            func.coalesce(func.sum(Order.total_amount), 0).label("total_spent"),
            func.max(Order.created_at).label("last_order_at"),
        )
        .group_by(Order.user_id)
        .subquery()
    )
    rows = (
        db.query(User, order_stats.c.order_count, order_stats.c.total_spent, order_stats.c.last_order_at)
        .outerjoin(order_stats, User.id == order_stats.c.user_id)
        .order_by(User.created_at.desc())
        .all()
    )
    addr_by_user: dict[int, list[Address]] = {}
    for a in db.query(Address).order_by(Address.is_default.desc(), Address.created_at.desc()).all():
        addr_by_user.setdefault(a.user_id, []).append(a)
    return [
        {
            "user": u,
            "order_count": int(oc or 0),
            "total_spent": float(ts or 0.0),
            "last_order_at": lo,
            "addresses": addr_by_user.get(u.id, []),
        }
        for (u, oc, ts, lo) in rows
    ]


@router.get("/admin/users")
def admin_list_users(db: Session = Depends(get_db), _=Depends(get_current_admin)):
    return [
        {
            "id": r["user"].id,
            "name": r["user"].name,
            "email": r["user"].email,
            "phone": r["user"].phone,
            "is_admin": r["user"].is_admin,
            "is_active": r["user"].is_active,
            "email_verified": r["user"].email_verified,
            "created_at": r["user"].created_at,
            "order_count": r["order_count"],
            "total_spent": r["total_spent"],
            "last_order_at": r["last_order_at"],
            "addresses": [
                {
                    "id": a.id,
                    "full_name": a.full_name,
                    "phone": a.phone,
                    "address_line1": a.address_line1,
                    "address_line2": a.address_line2,
                    "city": a.city,
                    "state": a.state,
                    "pincode": a.pincode,
                    "is_default": a.is_default,
                }
                for a in r["addresses"]
            ],
        }
        for r in _user_rows(db)
    ]


@router.get("/admin/users/export")
def admin_export_users(db: Session = Depends(get_db), _=Depends(get_current_admin)):
    """Return an .xlsx workbook with a Users sheet and an Addresses sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = _user_rows(db)
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0B2340")
    header_align = Alignment(horizontal="left", vertical="center")

    def _write_header(sheet, cols):
        for i, col in enumerate(cols, start=1):
            c = sheet.cell(row=1, column=i, value=col)
            c.font = header_font
            c.fill = header_fill
            c.alignment = header_align
        sheet.freeze_panes = "A2"

    def _autosize(sheet):
        for col in sheet.columns:
            length = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            sheet.column_dimensions[col[0].column_letter].width = min(length + 2, 60)

    users_sheet = wb.active
    users_sheet.title = "Users"
    user_cols = [
        "ID", "Name", "Email", "Phone", "Email Verified", "Admin", "Active",
        "Signed Up", "Orders", "Total Spent (₹)", "Last Order",
        "Default Address", "City", "State", "Pincode",
    ]
    _write_header(users_sheet, user_cols)

    def _fmt_dt(dt):
        if not dt:
            return ""
        return dt.strftime("%Y-%m-%d %H:%M") if isinstance(dt, datetime) else str(dt)

    for i, r in enumerate(rows, start=2):
        u = r["user"]
        default_addr = next((a for a in r["addresses"] if a.is_default), r["addresses"][0] if r["addresses"] else None)
        users_sheet.append([
            u.id,
            u.name,
            u.email,
            u.phone or "",
            "Yes" if u.email_verified else "No",
            "Yes" if u.is_admin else "No",
            "Yes" if u.is_active else "No",
            _fmt_dt(u.created_at),
            r["order_count"],
            round(r["total_spent"], 2),
            _fmt_dt(r["last_order_at"]),
            (default_addr.address_line1 + (", " + default_addr.address_line2 if default_addr.address_line2 else "")) if default_addr else "",
            default_addr.city if default_addr else "",
            default_addr.state if default_addr else "",
            default_addr.pincode if default_addr else "",
        ])
    _autosize(users_sheet)

    addr_sheet = wb.create_sheet("Addresses")
    _write_header(addr_sheet, [
        "User ID", "User Name", "User Email", "Recipient", "Phone",
        "Address Line 1", "Address Line 2", "City", "State", "Pincode", "Default",
    ])
    for r in rows:
        u = r["user"]
        for a in r["addresses"]:
            addr_sheet.append([
                u.id, u.name, u.email,
                a.full_name, a.phone,
                a.address_line1, a.address_line2 or "",
                a.city, a.state, a.pincode,
                "Yes" if a.is_default else "No",
            ])
    _autosize(addr_sheet)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"chhatak-users-{datetime.utcnow().strftime('%Y%m%d-%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
