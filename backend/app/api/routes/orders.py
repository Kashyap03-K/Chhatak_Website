import io
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.orm import Session

from sqlalchemy import func

from app.services.invoice import generate_invoice_pdf
from app.services.email import send_order_confirmation

from app.core.database import get_db
from app.api.deps import get_current_user, get_current_admin
from app.models.user import User
from app.models.order import Order, OrderItem
from app.models.cart import CartItem
from app.models.product import Product
from app.models.address import Address
from app.schemas.order import OrderOut, OrderCreate, OrderStatusUpdate

router = APIRouter(prefix="/orders", tags=["orders"])


@router.get("/", response_model=list[OrderOut])
def list_orders(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    return db.query(Order).filter(Order.user_id == user.id).order_by(Order.created_at.desc()).all()


@router.get("/{order_id}", response_model=OrderOut)
def get_order(order_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    order = db.query(Order).filter(Order.id == order_id, Order.user_id == user.id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    return order


def _build_order_from_cart(user: User, shipping_address: str, payment_method: str, db: Session) -> Order:
    cart_items = db.query(CartItem).filter(CartItem.user_id == user.id).all()
    if not cart_items:
        raise HTTPException(status_code=400, detail="Cart is empty")

    total = 0.0
    order_items: list[OrderItem] = []
    for ci in cart_items:
        product = db.query(Product).filter(Product.id == ci.product_id).first()
        if not product or not product.is_active:
            raise HTTPException(status_code=400, detail=f"Product {ci.product_id} unavailable")
        if product.stock < ci.quantity:
            raise HTTPException(status_code=400, detail=f"{product.name} has only {product.stock} left in stock")
        line_total = product.price * ci.quantity
        total += line_total
        order_items.append(OrderItem(
            product_id=product.id,
            quantity=ci.quantity,
            unit_price=product.price,
            total_price=line_total,
        ))

    order = Order(
        user_id=user.id,
        total_amount=total,
        shipping_address=shipping_address,
        status="pending_payment",
        payment_method=payment_method,
    )
    order.items = order_items
    db.add(order)
    return order


def _finalize_order(order: Order, db: Session) -> None:
    """Confirm order: decrement stock, clear cart, send confirmation email."""
    order.status = "confirmed"
    for item in order.items:
        product = db.query(Product).filter(Product.id == item.product_id).first()
        if product:
            product.stock = max(0, product.stock - item.quantity)
    db.query(CartItem).filter(CartItem.user_id == order.user_id).delete()
    db.commit()
    try:
        pdf = generate_invoice_pdf(order)
        send_order_confirmation(order, pdf)
    except Exception:
        pass


@router.post("/", response_model=OrderOut, status_code=status.HTTP_201_CREATED)
def create_order(body: OrderCreate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    pending = db.query(Order).filter(
        Order.user_id == user.id, Order.status == "pending_payment"
    ).first()
    if pending:
        pending.shipping_address = body.shipping_address
        pending.payment_method = body.payment_method
        db.commit()
        db.refresh(pending)
        return pending

    order = _build_order_from_cart(user, body.shipping_address, body.payment_method, db)
    db.commit()
    db.refresh(order)
    return order


@router.post("/place-cod", response_model=OrderOut, status_code=status.HTTP_201_CREATED)
def place_cod_order(body: OrderCreate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Create and immediately confirm a Cash-on-Delivery order."""
    order = _build_order_from_cart(user, body.shipping_address, "cod", db)
    db.commit()
    db.refresh(order)
    _finalize_order(order, db)
    db.refresh(order)
    return order


@router.get("/{order_id}/invoice")
def download_invoice(order_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    order = db.query(Order).filter(Order.id == order_id, Order.user_id == user.id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.status == "pending_payment":
        raise HTTPException(status_code=400, detail="Invoice available after payment")
    pdf = generate_invoice_pdf(order)
    return Response(
        content=pdf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=chhatak-invoice-{order.id}.pdf"},
    )


@router.get("/admin/all")
def admin_list_orders(db: Session = Depends(get_db), _=Depends(get_current_admin)):
    """Every order with the customer's profile, saved addresses, and lifetime stats."""
    orders = db.query(Order).order_by(Order.created_at.desc()).all()

    stats_rows = (
        db.query(
            Order.user_id,
            func.count(Order.id).label("order_count"),
            func.coalesce(func.sum(Order.total_amount), 0).label("total_spent"),
            func.max(Order.created_at).label("last_order_at"),
        )
        .group_by(Order.user_id)
        .all()
    )
    stats_by_user = {
        uid: {"order_count": int(oc or 0), "total_spent": float(ts or 0.0), "last_order_at": lo}
        for (uid, oc, ts, lo) in stats_rows
    }

    addr_by_user: dict[int, list[Address]] = {}
    for a in db.query(Address).order_by(Address.is_default.desc(), Address.created_at.desc()).all():
        addr_by_user.setdefault(a.user_id, []).append(a)

    def serialize_addr(a: Address):
        return {
            "id": a.id,
            "full_name": a.full_name,
            "phone": a.phone,
            "address_line1": a.address_line1,
            "address_line2": a.address_line2,
            "city": a.city,
            "state": a.state,
            "pincode": a.pincode,
            "is_default": a.is_default,
        }

    result = []
    for o in orders:
        u = o.user
        s = stats_by_user.get(o.user_id, {"order_count": 0, "total_spent": 0.0, "last_order_at": None})
        result.append({
            "id": o.id,
            "user_id": o.user_id,
            "total_amount": o.total_amount,
            "status": o.status,
            "payment_method": o.payment_method,
            "shipping_address": o.shipping_address,
            "created_at": o.created_at,
            "items": [
                {
                    "id": it.id,
                    "product_id": it.product_id,
                    "quantity": it.quantity,
                    "unit_price": it.unit_price,
                    "total_price": it.total_price,
                    "product": {
                        "id": it.product.id,
                        "name": it.product.name,
                        "slug": it.product.slug,
                        "flavor": getattr(it.product, "flavor", None),
                        "price": it.product.price,
                        "image_url": getattr(it.product, "image_url", None),
                    } if it.product else None,
                }
                for it in o.items
            ],
            "customer": {
                "id": u.id,
                "name": u.name,
                "email": u.email,
                "phone": u.phone,
                "is_admin": u.is_admin,
                "email_verified": u.email_verified,
                "created_at": u.created_at,
                "order_count": s["order_count"],
                "total_spent": s["total_spent"],
                "last_order_at": s["last_order_at"],
                "addresses": [serialize_addr(a) for a in addr_by_user.get(u.id, [])],
            } if u else None,
        })
    return result


@router.get("/admin/stats")
def admin_stats(db: Session = Depends(get_db), _=Depends(get_current_admin)):
    total_orders = db.query(func.count(Order.id)).scalar()
    total_revenue = db.query(func.coalesce(func.sum(Order.total_amount), 0)).filter(Order.status == "confirmed").scalar()
    pending_orders = db.query(func.count(Order.id)).filter(Order.status == "pending_payment").scalar()
    total_users = db.query(func.count(User.id)).scalar()
    total_products = db.query(func.count(Product.id)).filter(Product.is_active == True).scalar()
    return {
        "total_orders": total_orders,
        "total_revenue": float(total_revenue),
        "pending_orders": pending_orders,
        "total_users": total_users,
        "total_products": total_products,
    }


STATUS_LABELS = {
    "pending_payment": "Awaiting Payment",
    "confirmed": "Confirmed",
    "processing": "Processing",
    "shipped": "Shipped",
    "delivered": "Delivered",
    "cancelled": "Cancelled",
}


@router.get("/admin/export")
def admin_export_orders(db: Session = Depends(get_db), _=Depends(get_current_admin)):
    """Return an .xlsx workbook with Orders, Line Items, and Customers sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    orders = db.query(Order).order_by(Order.created_at.desc()).all()

    stats_rows = (
        db.query(
            Order.user_id,
            func.count(Order.id).label("order_count"),
            func.coalesce(func.sum(Order.total_amount), 0).label("total_spent"),
            func.max(Order.created_at).label("last_order_at"),
        )
        .group_by(Order.user_id)
        .all()
    )
    stats_by_user = {uid: (int(oc or 0), float(ts or 0.0), lo) for (uid, oc, ts, lo) in stats_rows}

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0B2340")
    header_align = Alignment(horizontal="left", vertical="center")

    def _write_header(sheet, cols):
        for i, col in enumerate(cols, start=1):
            c = sheet.cell(row=1, column=i, value=col)
            c.font = header_font
            c.fill = header_fill
            c.alignment = header_align
        sheet.freeze_panes = "A2"

    def _autosize(sheet):
        for col in sheet.columns:
            length = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            sheet.column_dimensions[col[0].column_letter].width = min(length + 2, 60)

    def _fmt_dt(dt):
        if not dt:
            return ""
        return dt.strftime("%Y-%m-%d %H:%M") if isinstance(dt, datetime) else str(dt)

    orders_sheet = wb.active
    orders_sheet.title = "Orders"
    _write_header(orders_sheet, [
        "Order ID", "Placed On", "Status", "Payment Method",
        "Items", "Total (₹)",
        "Customer Name", "Customer Email", "Customer Phone",
        "Lifetime Orders", "Lifetime Spent (₹)",
        "Shipping Address",
    ])
    for o in orders:
        u = o.user
        oc, ts, _ = stats_by_user.get(o.user_id, (0, 0.0, None))
        items_summary = "; ".join(
            f"{(it.product.name if it.product else 'Item')} × {it.quantity}" for it in o.items
        )
        orders_sheet.append([
            o.id,
            _fmt_dt(o.created_at),
            STATUS_LABELS.get(o.status, o.status),
            "Cash on Delivery" if o.payment_method == "cod" else "Razorpay",
            items_summary,
            round(o.total_amount, 2),
            u.name if u else "",
            u.email if u else "",
            (u.phone if u else "") or "",
            oc,
            round(ts, 2),
            (o.shipping_address or "").replace("\n", " | "),
        ])
    _autosize(orders_sheet)

    items_sheet = wb.create_sheet("Line Items")
    _write_header(items_sheet, [
        "Order ID", "Status", "Placed On", "Customer",
        "Product", "Quantity", "Unit Price (₹)", "Line Total (₹)",
    ])
    for o in orders:
        for it in o.items:
            items_sheet.append([
                o.id,
                STATUS_LABELS.get(o.status, o.status),
                _fmt_dt(o.created_at),
                o.user.name if o.user else "",
                it.product.name if it.product else f"Product #{it.product_id}",
                it.quantity,
                round(it.unit_price, 2),
                round(it.total_price, 2),
            ])
    _autosize(items_sheet)

    customers_sheet = wb.create_sheet("Customers")
    _write_header(customers_sheet, [
        "User ID", "Name", "Email", "Phone", "Email Verified",
        "Signed Up", "Lifetime Orders", "Lifetime Spent (₹)", "Last Order",
    ])
    seen: set[int] = set()
    for o in orders:
        u = o.user
        if not u or u.id in seen:
            continue
        seen.add(u.id)
        oc, ts, lo = stats_by_user.get(u.id, (0, 0.0, None))
        customers_sheet.append([
            u.id, u.name, u.email, u.phone or "",
            "Yes" if u.email_verified else "No",
            _fmt_dt(u.created_at),
            oc, round(ts, 2), _fmt_dt(lo),
        ])
    _autosize(customers_sheet)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"chhatak-orders-{datetime.utcnow().strftime('%Y%m%d-%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.patch("/{order_id}/status", response_model=OrderOut)
def update_order_status(order_id: int, body: OrderStatusUpdate, db: Session = Depends(get_db), _=Depends(get_current_admin)):
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    order.status = body.status
    db.commit()
    db.refresh(order)
    return order
